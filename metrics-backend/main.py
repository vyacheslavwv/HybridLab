from fastapi import FastAPI, HTTPException, status
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import logging

app = FastAPI(title="Metrics Data Logger", version="1.0.0")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

LOGS_DIR = Path("./logs")
LOGS_DIR.mkdir(exist_ok=True)
EXCEL_FILE = LOGS_DIR / "metrics_log.xlsx"


class MetricsData(BaseModel):
    pid_0d: float = Field(..., ge=0, le=300, description="Скорость автомобиля (км/ч)")
    pid_0c: float = Field(..., ge=0, le=8000, description="Обороты ДВС (об/мин)")
    pid_5c: float = Field(..., ge=-40, le=150, description="Температура масла (°C)")
    hv_soc: float = Field(..., ge=0, le=100, description="SOC HV батареи (%)")
    hv_volt: float = Field(..., ge=0, le=400, description="Напряжение HV шины (V)")
    electric_motor_torque: float = Field(..., ge=-500, le=500, description="Момент эл. мотора (Nm)")


def init_excel_file():
    """Инициализирует Excel файл с заголовками."""
    if not EXCEL_FILE.exists():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Metrics Log"
        
        headers = [
            "Timestamp",
            "Скорость (км/ч) [PID_0D]",
            "Обороты ДВС (об/мин) [PID_0C]",
            "Температура масла (°C) [PID_5C]",
            "SOC HV батареи (%) [HV_SOC]",
            "Напряжение HV шины (V) [HV_VOLT]",
            "Момент эл. мотора (Nm)"
        ]
        
        ws.append(headers)
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        ws.column_dimensions['A'].width = 22
        for col in ['B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 20
        
        wb.save(EXCEL_FILE)
        logger.info(f"Excel файл инициализирован: {EXCEL_FILE}")


def log_to_excel(data: MetricsData):
    """Логирует метрики в Excel файл."""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
        
        row_data = [
            timestamp,
            data.pid_0d,
            data.pid_0c,
            data.pid_5c,
            data.hv_soc,
            data.hv_volt,
            data.electric_motor_torque
        ]
        
        ws.append(row_data)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        last_row = ws.max_row
        for col in range(1, 8):
            cell = ws.cell(row=last_row, column=col)
            cell.border = border
            if col > 1:
                cell.alignment = Alignment(horizontal='center')
        
        wb.save(EXCEL_FILE)
        logger.info(f"Метрики записаны в строку {last_row}: {row_data}")
        
    except Exception as e:
        logger.error(f"Ошибка при записи в Excel: {str(e)}")
        raise


@app.on_event("startup")
async def startup_event():
    """Инициализирует Excel файл при запуске приложения."""
    init_excel_file()
    logger.info("Приложение запущено. Готово к приёму метрик.")


@app.get("/", tags=["Info"])
async def root():
    """Информация о сервисе."""
    return {
        "service": "Metrics Data Logger",
        "version": "1.0.0",
        "status": "running",
        "endpoints": {
            "POST /metrics": "Отправить метрики",
            "GET /status": "Статус сервиса",
            "GET /logs": "Получить статистику логов"
        }
    }


@app.post("/metrics", status_code=status.HTTP_201_CREATED, tags=["Metrics"])
async def receive_metrics(data: MetricsData):
    """
    Принимает метрики от фронтенда и логирует их в Excel.
    
    Все поля обязательны и должны быть в допустимых диапазонах.
    """
    try:
        log_to_excel(data)
        return {
            "status": "success",
            "message": "Метрики успешно зарегистрированы",
            "timestamp": datetime.now().isoformat(),
            "data_received": data.dict()
        }
    except Exception as e:
        logger.error(f"Ошибка при обработке метрик: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Ошибка при обработке метрик: {str(e)}"
        )


@app.get("/status", tags=["Info"])
async def status_check():
    """Проверка статуса сервиса и файла логов."""
    return {
        "service_status": "running",
        "excel_file_exists": EXCEL_FILE.exists(),
        "excel_file_path": str(EXCEL_FILE),
        "logs_directory": str(LOGS_DIR)
    }


@app.get("/logs", tags=["Info"])
async def get_logs_info():
    """Получить информацию о логированных метриках."""
    try:
        if not EXCEL_FILE.exists():
            return {"message": "Логов ещё нет", "total_records": 0}
        
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        
        total_records = ws.max_row - 1
        
        return {
            "total_records": total_records,
            "excel_file": str(EXCEL_FILE),
            "last_updated": EXCEL_FILE.stat().st_mtime,
            "message": f"Всего записей: {total_records}"
        }
    except Exception as e:
        logger.error(f"Ошибка при получении информации о логах: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(e)
        )


@app.get("/health", tags=["Info"])
async def health_check():
    """Liveness check для Docker."""
    return {"status": "healthy"}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)